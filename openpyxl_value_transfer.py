import openpyxl

# Load the source and destination workbooks
source_workbook = openpyxl.load_workbook('source.xlsx')
dest_workbook = openpyxl.load_workbook('destination.xlsx')

# Select the source and destination sheets
source_sheet = source_workbook['Sheet1']
dest_sheet = dest_workbook['Sheet2']

# Loop through the cells and copy their values to the destination sheet
for row in source_sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3):
    for cell in row:
        dest_sheet[cell.coordinate] = cell.value

# Save the changes to the destination workbook
dest_workbook.save('destination.xlsx')
